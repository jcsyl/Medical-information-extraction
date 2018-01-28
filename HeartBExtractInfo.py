import xlrd
import re
import openpyxl

def ExtractIndexInfo(inputstr):#提取指标 输入：描述
    inputstr = inputstr.strip()
    KeyVarible = ['AO根部', 'LA', 'RV常规', 'IVS', 'LVDd', 'LVDs', 'LVPW', 'RA横径', 'LVEF', 'AV', 'MVE', 'MVA']
    #print(inputstr)
    VaribleDic = {}
    for item in KeyVarible:
        content = re.findall(u'%s([\W\s\d\w]*?)[，：。]' % (item), inputstr)
        if len(content) > 0:
            VaribleDic[item] = content[0]
        else:
            VaribleDic[item] = 'None'
        #print(item,content)
    print(VaribleDic)
    return VaribleDic
def  Right_Ventricle(inputstr):#右室功能 输入：诊断
    Right_Ventricle_dict={'右室收缩功能':'不详','左室舒张功能':'不详'}
    return Right_Ventricle_dict
def  Left_Ventricle(inputstr):#左室功能 输入：诊断
    inputstr = inputstr.strip()
    Left_Ventricle_dic = {}
    des_dict = {0: '减退', 1: '未见异常', None:'不详'}
    str_list = ['减退','未见?异常']
    Systolic_flag = None
    Diastolic_flag = None
    Systolic_content = re.findall(u'(左心?室收缩功能[\W\s\d\w]*?[，：。、；])', inputstr)
    Diastolic_content = re.findall(u'左心?室[\W\s\d\w]*?(舒张功能[\W\s\d\w]*?[，：。、；])', inputstr)
    if len(Systolic_content) > 0:
        #print(Systolic_content)
        for index,item in enumerate(str_list):
            sub_content = re.findall(u'%s' % (item), Systolic_content[0])
            if len(sub_content) > 0:
                Systolic_flag = index
                break
    if len(Diastolic_content) > 0:
        #print(Diastolic_content)
        for index_one, item_one in enumerate(str_list):
            sub_content_one = re.findall(u'%s' % (item_one), Diastolic_content[0])
            if len(sub_content_one) > 0:
                Diastolic_flag = index_one
                break
    #print(Systolic_flag, Diastolic_flag)
    Left_Ventricle_dic['左室收缩功能'] = des_dict[Systolic_flag]
    Left_Ventricle_dic['左室舒张功能'] = des_dict[Diastolic_flag]
    #return Systolic_flag, Diastolic_flag
    print(Left_Ventricle_dic)
    return Left_Ventricle_dic

def AortaRoot(inputstr):#主动脉根部异常 输入：描述
    inputstr = inputstr.strip()
    AortaRoot_dict = {}
    root_dict = {0:'无',1:'有', None:'不详'}
    des_dict = {0:'否',1:'是', None:'不详'}
    AortaRoot_content = re.findall(u'主动脉(?:根部)?[^瓣][\W\s\d\w]*?(?:[^壁瓣腔]搏动[\W\s\d\w]*?)?[，：。、；]',inputstr)
    root_des = ['不宽','([增稍]?增?宽)|不好']
    pulse_des = ['尚?[好可]','([稍减]?低平?)|呈弓背样']
    root_flag = None
    pulse_flag = None
    abnormal = None
    if len(AortaRoot_content) > 0:
        for index_r,item_r in enumerate(root_des):
            sub_root_des = re.findall(u'主动脉(?:根部)?[^瓣][\W\s\d\w]*?%s' % (item_r), AortaRoot_content[0])
            if len(sub_root_des) > 0:
                root_flag = index_r
                break
        for index_p, item_p in enumerate(pulse_des):
            sub_pulse_des = re.findall(u'搏动[\W\s\d\w]*?%s' % (item_p), AortaRoot_content[0])
            if len(sub_pulse_des) > 0:
                pulse_flag = index_p
                break
        if root_flag == 0 and pulse_flag == 0:
            abnormal = 0
        if root_flag == 1 or pulse_flag == 1:
            abnormal = 1
        #print(AortaRoot_content,abnormal,root_flag,  pulse_flag)
    else:
        print('未找到相关字段')
    AortaRoot_dict['主动脉根部异常'] = root_dict[abnormal]
    AortaRoot_dict['宽度异常'] = des_dict[root_flag]
    AortaRoot_dict['搏动异常'] = des_dict[pulse_flag]
    print(AortaRoot_dict)
    #return root_flag, pulse_flag, abnormal
    return AortaRoot_dict
####################################################################################
#       Aortic_Vaives() 主动脉瓣开放异常关闭判断：返回 无 ，有  不详
#       Mitral_valve（）二尖瓣开放异常关闭判断 ：返回 无，有，不详
#       Tricuspid_value() 二尖瓣开放异常关闭判断：返回 无，有，不详
#       Atrial_Ventricular_Septa() 房室间隔有中断： 返回 无， 有， 不详
#       Main_pulmonary_artery() 主肺动脉异常判断：返回 无，有，不详
#       Pasp() 估测肺动脉收缩压: 返回 数值，none
#                                                    author：gsc  date:2017/12/26
###################################################################################
def  Aortic_Vaives(inputstr):
    VaribleDic={}
    inputstr = inputstr.strip()
    aortic_dict = {0:'无',1:'有', 2:'不详'}
    aortic_dos=['尚?[好可]']
    Aortic_Vaives_content = re.findall(u'主动脉瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?[，：。、；]',inputstr)
    Aortic_Vaives_content2 = re.findall(u'各瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?[，：。、；]',inputstr)
    if len(Aortic_Vaives_content) > 0:
        sub_aortic_des = re.findall(u'主动脉瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?%s' % (aortic_dos[0]), Aortic_Vaives_content[0])
        if len(sub_aortic_des)>0:
            VaribleDic['主动脉瓣开放关闭异常']=aortic_dict[0]
            print(VaribleDic)
            return VaribleDic
    if len(Aortic_Vaives_content2)>0:
        VaribleDic['主动脉瓣开放关闭异常']=aortic_dict[0]
        print(VaribleDic)
        return VaribleDic
    else:
        VaribleDic['主动脉瓣开放关闭异常']=aortic_dict[2]
        print(VaribleDic)
        return VaribleDic

def Mitral_valve(inputstr):
    VaribleDic={}
    inputstr = inputstr.strip()
    mitral_dict = {0:'无',1:'有', 2:'不详'}
    mitral_dos=['尚?[好可]']
    mitral_Valve_content = re.findall(u'二尖瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?[，：。、；]',inputstr)
    mitral_Valves_content2 = re.findall(u'各瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?[，：。、；]',inputstr)
    if len(mitral_Valve_content) > 0:
        sub_mitral_des = re.findall(u'二尖瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?%s' % (mitral_dos[0]), mitral_Valve_content[0])
        if len(sub_mitral_des)>0:
            VaribleDic['二尖瓣开放关闭异常']=mitral_dict[0]
            print(VaribleDic)
            return VaribleDic
    if len(mitral_Valves_content2)>0:
        VaribleDic['二尖瓣开放关闭异常']=mitral_dict[0]
        print(VaribleDic)
        return VaribleDic
    else:
        VaribleDic['二尖瓣开放关闭异常']=mitral_dict[2]
        print(VaribleDic)
        return VaribleDic

def Tricuspid_value(inputstr):
    VaribleDic={}
    inputstr = inputstr.strip()
    tricuspid_dict = {0:'无',1:'有', 2:'不详'}
    tricuspid_dos=['尚?[好可]']
    tricuspid_Valve_content = re.findall(u'三尖瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?[，：。、；]',inputstr)
    tricuspid_Valves_content2 = re.findall(u'各瓣膜[\W\s\d\w]*?开放关闭[\W\s\d\w]*?[，：。、；]',inputstr)
    if len(tricuspid_Valve_content) > 0:
        sub_tricuspid_des = re.findall(u'三尖瓣[\W\s\d\w]*?开放关闭[\W\s\d\w]*?%s' % (tricuspid_dos[0]), tricuspid_Valve_content[0])
        if len(sub_tricuspid_des)>0:
            VaribleDic['三尖瓣开放关闭异常']=tricuspid_dict[0]
            print(VaribleDic)
            return VaribleDic
    if len(tricuspid_Valves_content2)>0:
        VaribleDic['三尖瓣开放关闭异常']=tricuspid_dict[0]
        print(VaribleDic)
        return VaribleDic
    else:
        VaribleDic['三尖瓣开放关闭异常']=tricuspid_dict[2]
        print(VaribleDic)
        return VaribleDic

def Atrial_Ventricular_Septa(inputstr):
    VaribleDic={}
    inputstr = inputstr.strip()
    atrial_ventricular_dict = {0:'无',1:'有',2:'不详'}
    atrial_ventricular_dos=['未见']
    atrial_ventricular_content = re.findall(u'房[，：。、；]室间隔[\W\s\d\w]*?[，：。、；]',inputstr)
    if len(atrial_ventricular_content)>0:
        sub_a_v_content = re.findall(u'%s' % (atrial_ventricular_dos[0]),atrial_ventricular_content[0])
        if len(sub_a_v_content)>0:
            VaribleDic['房、室间隔有中断']=atrial_ventricular_dict[0]
            print(VaribleDic)
            return VaribleDic
    else:
        VaribleDic['房、室间隔有中断']=atrial_ventricular_dict[2]
        print(VaribleDic)
        return VaribleDic

def Main_pulmonary_artery(inputstr):
    VaribleDic={}
    inputstr = inputstr.strip()
    main_pulmonary_dict={0:'无',1:'有',2:'不详'}
    main_pulmonary_dos=['未见增宽']
    main_pulmonary_content = re.findall(u'(?:主)?肺动脉[\W\s\d\w]*?[，：。、；]',inputstr)
    if len(main_pulmonary_content)>0:
        sub_m_p_content = re.findall(u'(?:主)肺动脉%s' %(main_pulmonary_dos[0]),main_pulmonary_content[0])
        if len(sub_m_p_content)>0:
            VaribleDic['主肺动脉异常']=main_pulmonary_dict[0]
            print(VaribleDic)
            return VaribleDic
        else:
            VaribleDic['主肺动脉异常']=main_pulmonary_dict[2]
            print(VaribleDic)
            return VaribleDic
    else:
        VaribleDic['主肺动脉异常']=main_pulmonary_dict[2]
        return VaribleDic

def Pasp(inputstr):
    inputstr = inputstr.strip()
    KeyVarible = ['估测肺动脉收缩压', 'PASP']
    #print(inputstr)
    VaribleDic = {}
    for item in KeyVarible:
        content = re.findall(u'%s[）]?([\W\s\d\w]*?)[，：。]' % (item), inputstr)
        if len(content) > 0:
            VaribleDic['估测肺动脉收缩压'] = content[0]
        else:
            VaribleDic['估测肺动脉收缩压'] = 'None'
        #print(item,content)
    print(VaribleDic)
    return VaribleDic

def find_word(sentence, word):
    if sentence.strip().find(word) == -1:
        return 0
    else:
        return 1

def s2list(s): # 从report discribe中提取
    if isinstance(s, str):
        tmps = s.split('\n')
        return tmps
    elif isinstance(s, list):
        return s
    else:
        print('Error!!')

def fls(discribeS):
    result = {
        '二尖瓣':'',
        '三尖瓣':'',
        '主动脉瓣': '',
        '肺动脉瓣': ''
    }
    discribeS = s2list(discribeS)
    flsmjp = re.compile(r'返流束面积(.+cm2)')
    print(flsmjp)
    for s in discribeS:
        if find_word(s, '返流束') or find_word(s,'局限返流'):
            if find_word(s, '主动脉瓣'):
                result['主动脉瓣'] = '有'
                if re.findall(flsmjp, s):
                    result['主动脉瓣'] = re.findall(flsmjp, s)[0].strip()
            if find_word(s, '肺动脉瓣') :
                result['肺动脉瓣'] = '有'
                if re.findall(flsmjp, s):
                    result['肺动脉瓣'] = re.findall(flsmjp, s)[0].strip()
            if find_word(s, '二尖瓣') :
                result['二尖瓣'] = '有'
               
                if re.findall(flsmjp, s):
                    result['二尖瓣'] = re.findall(flsmjp, s)[0].strip()
            if find_word(s, '三尖瓣') :
                result['三尖瓣'] = '有'
                if re.findall(flsmjp, s):
                    result['三尖瓣'] = re.findall(flsmjp, s)[0].strip()
            if find_word(s, '二、三尖瓣') :
                result['三尖瓣'] = '有'
                result['二尖瓣'] = '有'
    print(result)
    return result
################################################################################################################
##   function1()：该函数用于检测异常与否 
##              需要添加的关键字 list1：异常的描述 list2:检测的部位  list3:返回诊断结果的格式
##   function2(): 该函数用来提出检测值
##              需要添加的关键字 list1:检测部位  list:返回诊断结果格式
##              
##                                                  author:gsc  date:2018/1/7
################################################################################################################
def  function1(inputstr):
    inputstr = inputstr.strip()
    print(inputstr)
    return_dict={}
    refer_dict = {0:'无',1:'有', 2:'不详'}
    describle_dict = ['开放关闭尚?[好可]','未见中断']  #z异常描述
    organ_list = ['二尖瓣','各瓣膜','三尖瓣','主动脉瓣','房[，：。、；]?室间隔'] #检测部位
    return_list = ['主动脉瓣开放关闭异常','二尖瓣开放关闭异常','三尖瓣开放关闭异常','房、室间隔有中断',''] #诊断结果
    for item in return_list:
        return_dict[item]='不详'
    for item_organ in organ_list:
        for item_des in describle_dict:
            values_content = re.findall(u'%s[\W\s\d\w]*?%s[\W\s\d\w]*?[，：。、；]'%(item_organ,item_des),inputstr)
            if  len(values_content)>0:

                if item_organ=='主动脉瓣' and item_des=='开放关闭尚?[好可]':
                    return_dict['主动脉瓣开放关闭异常'] = refer_dict[0]
                if item_organ=='二尖瓣' and item_des=='开放关闭尚?[好可]':
                    return_dict['二尖瓣开放关闭异常'] = refer_dict[0]
                if item_organ=='三尖瓣' and item_des=='开放关闭尚?[好可]':
                    return_dict['三尖瓣开放关闭异常'] = refer_dict[0]
                if item_organ=='房[，：。、；]?室间隔' and item_des=='未见中断':
                    return_dict['房、室间隔有中断'] = refer_dict[0]
                if item_organ=='各瓣膜' and item_des=='开放关闭尚?[好可]':
                    return_dict['三尖瓣开放关闭异常'] = refer_dict[0]
                    return_dict['二尖瓣开放关闭异常'] = refer_dict[0]

    print(return_dict)
    return return_dict

def function2(inputstr):
    inputstr = inputstr.strip()
    measure_list = ['估测肺动脉收缩压','PASP']
    return_list=['估测肺动脉收缩压']
    return_dict={}
    for item in return_list:
        return_dict[item]='None'
    for item_measure in measure_list:
        values_content = re.findall(u'%s[）)：:]?([\W\s\d\w]*?)[，：。]' % (item_measure), inputstr)
        if len(values_content)>0:
            if item_measure=='估测肺动脉收缩压' or item_measure=='PASP':
                return_dict['估测肺动脉收缩压'] = values_content[0]
    print(return_dict)
    return return_dict

###############################################################################################################################


def Left_Heart_System(inputstr):#左心系统 输入：描述
    inputstr = inputstr.strip()
    Left_Heart_System_dict = {}
    Left_Heart_System_abnormal_des={0:'无',1:'有', None:'不详'}
    Left_Heart_System_other_des = {0:'否',1:'是', None:'不详'}
    Left_Heart_house_des=['(?:未见增大|不大)','稍?增?[大高]']
    Left_Heart_ventricle_des=['未见增大','稍?增?大']
    Left_Combustor_Wall_des=['未见增厚','变[薄厚]']
    Left_Combustor_Wall_pulse_des=['尚?[好可]','(?:明显|稍|普遍)?减弱']
    first_flag = 0
    Left_Heart_house_flag = None
    Left_Heart_ventricle_flag = None
    Left_Combustor_Wall_incrassation_flag = None
    Left_Combustor_Wall_pulse_flag = None
    all_Left_Combustor_Wall = None
    abnormal = None
    The_first_content = re.findall(u'各房、?室腔?不大',inputstr)
    if len(The_first_content) > 0:
        first_flag = 1
    Left_Heart_System_content = re.findall(u'(?:左房|左室)[^收缩功能][\W\s\d\w]*室壁[\W\s\d\w]*?[，：。、；]',inputstr)
    if len(Left_Heart_System_content) > 0:
        #print( Left_Heart_System_content)
        for index_h,item_h in enumerate( Left_Heart_house_des):
            Left_Heart_house_content = re.findall(u'左房[\W\s\d\w]*?腔?[\W\s\d\w]*?%s' % (item_h),Left_Heart_System_content[0])
            if len(Left_Heart_house_content) > 0:
                Left_Heart_house_flag = index_h
                #print(Left_Heart_house_content, Left_Heart_house_flag)
                break
        for index_v,item_v in enumerate(Left_Heart_ventricle_des):
            Left_Heart_ventricle_content =  re.findall(u'左室腔?%s' % (item_v),Left_Heart_System_content[0])
            if len(Left_Heart_ventricle_content) > 0:
                Left_Heart_ventricle_flag = index_v
                break
        for index_w, item_w in enumerate(Left_Combustor_Wall_des):
            Left_Combustor_Wall_content = re.findall(u'[左余]?[余左]?[^右]室[\W\s\d\w]*?壁[\W\s\d\w]*?%s' % (item_w),Left_Heart_System_content[0])
            if len(Left_Combustor_Wall_content) > 1:
                for sub_index, sub_item in enumerate(Left_Combustor_Wall_des):
                    sub_Left_Combustor_Wall_content = re.findall(u'[左余]?[余左]?[^右]室[\W\s\d\w]*?壁[\W*\s\d\w]*?%s' % (sub_index),
                                                                 Left_Combustor_Wall_content[1])
                    if len(sub_Left_Combustor_Wall_content) > 0:
                        Left_Combustor_Wall_incrassation_flag = sub_index
                        break
            if len(Left_Combustor_Wall_content) > 0 and len(Left_Combustor_Wall_content) < 2:
                Left_Combustor_Wall_incrassation_flag = index_w
                break
        for index_p, item_p in enumerate(Left_Combustor_Wall_pulse_des):
            Left_Combustor_Wall_pulse_content = re.findall(u'[左余]?[^右]室[\W\s\d\w]*?壁[\W\s\d\w]*?[运搏]动%s' % (item_p), Left_Heart_System_content[0])
            if len(Left_Combustor_Wall_pulse_content) > 0:
                Left_Combustor_Wall_pulse_flag = index_p
                break
    else:
       if first_flag == 1:
           Left_Heart_house_flag = 0
           Left_Heart_ventricle_flag = 0
           for index_w, item_w in enumerate(Left_Combustor_Wall_des):
               Left_Combustor_Wall_content = re.findall(u'[左余]?[余左]?[^右]室[\W\s\d\w]*?壁[\W\s\d\w]*?%s' % (item_w),
                                                       inputstr)
               if len(Left_Combustor_Wall_content) > 1:
                   for sub_index, sub_item in enumerate(Left_Combustor_Wall_des):
                       sub_Left_Combustor_Wall_content = re.findall(
                           u'[左余]?[余左]?[^右]室[\W\s\d\w]*?壁[\W*\s\d\w]*?%s' % (sub_index),
                           Left_Combustor_Wall_content[1])
                       if len(sub_Left_Combustor_Wall_content) > 0:
                           Left_Combustor_Wall_incrassation_flag = sub_index
                           break
               if len(Left_Combustor_Wall_content) > 0 and len(Left_Combustor_Wall_content) < 2:
                   Left_Combustor_Wall_incrassation_flag = index_w
                   break
           for index_p, item_p in enumerate(Left_Combustor_Wall_pulse_des):
               Left_Combustor_Wall_pulse_content = re.findall(u'[左余]?[^右]室[\W\s\d\w]*?壁[\W\s\d\w]*?[运搏]动%s' % (item_p),
                                                              inputstr)
               if len(Left_Combustor_Wall_pulse_content) > 0:
                   Left_Combustor_Wall_pulse_flag = index_p
                   break
    if Left_Combustor_Wall_pulse_flag == 1 or  Left_Combustor_Wall_incrassation_flag == 1:
        all_Left_Combustor_Wall = 1
    if Left_Combustor_Wall_pulse_flag == 0 and Left_Combustor_Wall_incrassation_flag == 0:
        all_Left_Combustor_Wall = 0
    if Left_Heart_house_flag == 0 and Left_Heart_ventricle_flag == 0 and all_Left_Combustor_Wall == 0:
        abnormal = 0
    if Left_Heart_house_flag == 1 or Left_Heart_ventricle_flag == 1 or all_Left_Combustor_Wall == 1:
        abnormal = 1
    #print(abnormal, Left_Heart_house_flag, Left_Heart_ventricle_flag,all_Left_Combustor_Wall)
    Left_Heart_System_dict['左心系统异常'] =  Left_Heart_System_abnormal_des[abnormal]
    Left_Heart_System_dict['左房大'] = Left_Heart_System_other_des[ Left_Heart_house_flag]
    Left_Heart_System_dict['左室大'] = Left_Heart_System_other_des[ Left_Heart_ventricle_flag]
    Left_Heart_System_dict['室壁异常'] = Left_Heart_System_other_des[ all_Left_Combustor_Wall]
    #return abnormal, Left_Heart_house_flag,Left_Heart_ventricle_flag,all_Left_Combustor_Wall
    print(Left_Heart_System_dict)
    return Left_Heart_System_dict

def Right_Heart_System(inputstr):#右心系统 输入：描述
    inputstr = inputstr.strip()
    Right_Heart_System_dict = {}
    Right_Heart_System_abnormal_des = {0: '无', 1: '有', None: '不详'}
    Right_Heart_System_other_des = {0: '否', 1: '是', None: '不详'}
    Right_Heart_house_des = ['(?:未见增大|不大)', '稍?增?[大高]']
    Right_Heart_ventricle_des = ['未见增大', '稍?增?大']
    Right_Heart_house_flag = None
    Right_Heart_ventricle_flag = None
    Right_Combustor_Wall_incrassation_flag = None
    abnormal = None
    The_first_content = re.findall(u'各房、?室腔?不大', inputstr)
    if len(The_first_content) > 0:
        Right_Heart_house_flag = 0
        Right_Heart_ventricle_flag = 0
    else:
        for index_h, item_h in enumerate(Right_Heart_house_des):
            Right_Heart_house_content = re.findall(u'右房[\W\s\d\w]*?腔?[\W\s\d\w]*?%s' % (item_h),
                                                   inputstr)
            if len(Right_Heart_house_content) > 0:
                Right_Heart_house_flag = index_h
                # print( Right_Heart_house_content, Right_Heart_house_flag )
                break
        for index_v, item_v in enumerate(Right_Heart_ventricle_des):
            Right_Heart_ventricle_content = re.findall(u'右室腔?%s' % (item_v), inputstr)
            if len(Right_Heart_ventricle_content) > 0:
                Right_Heart_ventricle_flag = index_v
                # print(Right_Heart_ventricle_content, Right_Heart_ventricle_flag)
                break

    if Right_Heart_house_flag == 0 and  Right_Heart_ventricle_flag == 0:
        abnormal = 0
    if Right_Heart_house_flag == 1 or  Right_Heart_ventricle_flag == 1:
        abnormal = 1
    Right_Heart_System_dict['右心系统异常'] = Right_Heart_System_abnormal_des[abnormal]
    Right_Heart_System_dict['右房大'] = Right_Heart_System_other_des[Right_Heart_house_flag]
    Right_Heart_System_dict['右室大'] = Right_Heart_System_other_des[Right_Heart_ventricle_flag]
    Right_Heart_System_dict['室壁异常'] = Right_Heart_System_other_des[Right_Combustor_Wall_incrassation_flag]
    print(Right_Heart_System_dict)
    return Right_Heart_System_dict
    #return Right_Heart_house_flag, Right_Heart_ventricle_flag

if __name__ == '__main__':
    wb = xlrd.open_workbook('心脏B超.xls')
    ws = wb.sheet_by_name('Sheet1')
    rows = ws.nrows
    cols = ws.ncols
    for row in range(3,4):
        PatientName = ws.cell(row, 0).value
        print(PatientName)
        ReportDescribe = ws.cell(row, 1).value
        ReportDiagnose = ws.cell(row, 2).value
        #MainReportDescribe = ReportDescribe+"\n"+ReportDiagnose
        #print(MainReportDescribe)
        # Var = ExtractIndexInfo(ReportDescribe)
        # Left_Systolic_flag, Left_Diastolic_flag = Left_Ventricle(ReportDiagnose)
        # AortaRoot(ReportDescribe)
        # Left_Heart_System(ReportDescribe)
        # Right_Heart_System(ReportDescribe)
        #Aortic_Vaives(ReportDescribe)
        #Var = ExtractInfo(MainReportDescribe)
        #Left_Systolic_flag, Left_Diastolic_flag = Left_Ventricle(MainReportDescribe)
        #root_flag, pulse_flag, abnormal = AortaRoot(MainReportDescribe)
        #print( Left_Systolic_flag,  Left_Diastolic_flag)
        # Aortic_Vaives(ReportDescribe)
        # Mitral_valve(ReportDescribe)
        # Tricuspid_value(ReportDescribe)
        #function1(ReportDescribe)
        # Atrial_Ventricular_Septa(ReportDescribe)
        # Main_pulmonary_artery(ReportDescribe)
        # Pasp(ReportDescribe)
        # function2(ReportDescribe)
        fls(ReportDescribe)






